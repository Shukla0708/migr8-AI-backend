import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from services.rules_engine import validate_cell

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MAX_STORED_EXCEPTIONS = 60


def extract_headers(file_bytes: bytes) -> list[str]:
    """Read only the header row — used right after upload to populate the
    'Validation Rules Configuration' UI with real column names."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    return [str(h).strip() for h in header_row if h is not None]


def run_validation(file_bytes: bytes, field_configs: list[dict]):
    """
    field_configs: one dict per field, matching ValidationField columns.

    Returns (annotated_workbook_bytes, stats_dict, exceptions_list).
    Output keeps the same layout/format as the source file, with:
      - failing cells filled red
      - one appended 'Validation_Failure_Reason' column containing the
        combined reasons ("<Field>: <reason>; ...") for every failing
        cell in that row.
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    name_to_col = {str(name).strip(): idx for idx, name in enumerate(header) if name}
    reason_col_idx = len(header)
    ws.cell(row=1, column=reason_col_idx + 1, value="Validation_Failure_Reason")

    cfg_by_field = {c["field_name"]: c for c in field_configs}
    seen_keys_by_field = {c["field_name"]: set() for c in field_configs if c["flag_key"]}

    total_rows = valid_rows = invalid_rows = total_errors = critical_errors = 0
    errors_by_field: dict[str, int] = {}
    errors_by_type: dict[str, int] = {}
    exceptions: list[dict] = []

    for row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue  # skip fully blank trailing rows

        total_rows += 1
        row_reasons: list[str] = []
        row_has_error = False

        for field_name, cfg in cfg_by_field.items():
            col_idx = name_to_col.get(field_name)
            if col_idx is None:
                continue
            cell = row[col_idx]
            seen_keys = seen_keys_by_field.get(field_name, set())
            reasons = validate_cell(cell.value, cfg, seen_keys)

            if reasons:
                row_has_error = True
                cell.fill = RED_FILL
                is_critical = cfg["flag_mandatory"] or cfg["flag_key"]

                for reason in reasons:
                    row_reasons.append(f"{field_name}: {reason}")
                    total_errors += 1
                    errors_by_field[field_name] = errors_by_field.get(field_name, 0) + 1
                    bucket = reason.split(" ")[0]
                    errors_by_type[bucket] = errors_by_type.get(bucket, 0) + 1
                    if is_critical:
                        critical_errors += 1

                    if len(exceptions) < MAX_STORED_EXCEPTIONS:
                        exceptions.append({
                            "row_number": row[0].row,
                            "field_name": field_name,
                            "actual_value": str(cell.value),
                            "expected_value": _expected_label(cfg),
                            "error_type": reason,
                            "severity": "error" if is_critical else "warning",
                        })

        if row_has_error:
            invalid_rows += 1
            ws.cell(row=row[0].row, column=reason_col_idx + 1, value="; ".join(row_reasons))
        else:
            valid_rows += 1

    out = io.BytesIO()
    wb.save(out)
    wb.close()

    health_score = round((valid_rows / total_rows) * 100, 2) if total_rows else 100.0

    # errorsByType is rendered as a donut chart on the frontend and expects a
    # hex color per slice plus a percentage-of-total value (not a raw count).
    PALETTE = ["#004da4", "#6063ee", "#8a3500", "#c2c6d5", "#0f9d58", "#d93025"]
    type_total = sum(errors_by_type.values()) or 1
    errors_by_type_chart = [
        {
            "label": label,
            "value": round((count / type_total) * 100),
            "color": PALETTE[i % len(PALETTE)],
        }
        for i, (label, count) in enumerate(errors_by_type.items())
    ]

    stats = {
        "total_records": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "total_errors": total_errors,
        "critical_errors": critical_errors,
        "health_score": health_score,
        "errors_by_field": [{"field": k, "count": v} for k, v in errors_by_field.items()],
        "errors_by_type": errors_by_type_chart,
    }
    return out.getvalue(), stats, exceptions


def _expected_label(cfg: dict) -> str:
    if cfg["flag_email"]:
        return "Valid email format"
    if cfg["flag_mobile"]:
        return "Valid mobile number"
    if cfg["flag_date"]:
        return "Valid date"
    if cfg["max_length"]:
        return f"Max length {cfg['max_length']}"
    return cfg["data_type"]
